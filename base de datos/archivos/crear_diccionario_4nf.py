from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

FILE_NAME = "17_diccionario_normalizado_4NF.xlsx"

headers = ["Columna", "Tipo", "Null", "Default", "Llave", "Relación", "Descripción"]

# Definición de tablas normalizadas a 4NF
tables = [
    {
        "name": "rol",
        "description": "Catálogo de roles del sistema.",
        "columns": [
            ("rol_id", "INT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador de rol"),
            ("clave", "VARCHAR(50)", "NN", "-", "UK", "-", "Clave única del rol"),
            ("nombre", "VARCHAR(100)", "NN", "-", "-", "-", "Nombre del rol"),
            ("descripcion", "VARCHAR(500)", "N", "-", "-", "-", "Descripción del rol"),
            ("activo", "BIT", "NN", "1", "-", "-", "Bandera de vigencia"),
            ("fecha_creacion", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha de creación"),
        ],
    },
    {
        "name": "usuario",
        "description": "Usuarios del sistema.",
        "columns": [
            ("usuario_id", "INT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador de usuario"),
            ("nombre_pila", "VARCHAR(100)", "NN", "-", "-", "-", "Nombre(s) de pila"),
            ("apellido_paterno", "VARCHAR(100)", "NN", "-", "-", "-", "Apellido paterno"),
            ("apellido_materno", "VARCHAR(100)", "N", "-", "-", "-", "Apellido materno"),
            ("correo_institucional", "VARCHAR(180)", "NN", "-", "UK", "-", "Correo institucional único"),
            ("puesto", "VARCHAR(120)", "N", "-", "-", "-", "Puesto o cargo"),
            ("activo", "BIT", "NN", "1", "-", "-", "Bandera de vigencia"),
            ("fecha_creacion", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha de creación"),
        ],
    },
    {
        "name": "usuario_rol",
        "description": "Asignación N:M entre usuarios y roles.",
        "columns": [
            ("usuario_rol_id", "INT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador técnico"),
            ("usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Usuario"),
            ("rol_id", "INT", "NN", "-", "FK", "rol.rol_id", "Rol"),
            ("activo", "BIT", "NN", "1", "-", "-", "Asignación vigente"),
            ("fecha_asignacion", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha de asignación"),
        ],
    },
    {
        "name": "area",
        "description": "Estructura jerárquica de áreas organizacionales.",
        "columns": [
            ("area_id", "INT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador de área"),
            ("area_padre_id", "INT", "N", "-", "FK", "area.area_id", "Área padre"),
            ("clave", "VARCHAR(40)", "NN", "-", "UK", "-", "Clave única de área"),
            ("nombre", "VARCHAR(150)", "NN", "-", "-", "-", "Nombre del área"),
            ("descripcion", "VARCHAR(500)", "N", "-", "-", "-", "Descripción"),
            ("activa", "BIT", "NN", "1", "-", "-", "Bandera de vigencia"),
        ],
    },
    {
        "name": "usuario_area",
        "description": "Asignación N:M entre usuarios y áreas.",
        "columns": [
            ("usuario_area_id", "INT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador técnico"),
            ("usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Usuario"),
            ("area_id", "INT", "NN", "-", "FK", "area.area_id", "Área"),
            ("es_principal", "BIT", "NN", "0", "-", "-", "Área principal del usuario"),
            ("fecha_asignacion", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha de asignación"),
        ],
    },
    {
        "name": "sistema",
        "description": "Catálogo de sistemas/aplicaciones.",
        "columns": [
            ("sistema_id", "INT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador de sistema"),
            ("clave", "VARCHAR(50)", "NN", "-", "UK", "-", "Clave única"),
            ("nombre", "VARCHAR(150)", "NN", "-", "-", "-", "Nombre del sistema"),
            ("descripcion", "VARCHAR(1000)", "N", "-", "-", "-", "Descripción"),
            ("activo", "BIT", "NN", "1", "-", "-", "Vigencia"),
        ],
    },
    {
        "name": "area_sistema",
        "description": "Relación N:M entre áreas y sistemas.",
        "columns": [
            ("area_sistema_id", "INT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador técnico"),
            ("area_id", "INT", "NN", "-", "FK", "area.area_id", "Área"),
            ("sistema_id", "INT", "NN", "-", "FK", "sistema.sistema_id", "Sistema"),
            ("fecha_alta", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha de alta"),
        ],
    },
    {
        "name": "tipo_solicitud",
        "description": "Catálogo de tipos de solicitud.",
        "columns": [
            ("tipo_solicitud_id", "INT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("clave", "VARCHAR(40)", "NN", "-", "UK", "-", "Clave única"),
            ("nombre", "VARCHAR(120)", "NN", "-", "-", "-", "Nombre"),
            ("activa", "BIT", "NN", "1", "-", "-", "Vigencia"),
        ],
    },
    {
        "name": "prioridad_solicitud",
        "description": "Catálogo de prioridades.",
        "columns": [
            ("prioridad_solicitud_id", "INT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("clave", "VARCHAR(30)", "NN", "-", "UK", "-", "Clave"),
            ("nombre", "VARCHAR(80)", "NN", "-", "-", "-", "Nombre"),
            ("orden", "INT", "NN", "-", "-", "-", "Orden de atención"),
            ("activa", "BIT", "NN", "1", "-", "-", "Vigencia"),
        ],
    },
    {
        "name": "estado_solicitud",
        "description": "Catálogo de estados de una solicitud.",
        "columns": [
            ("estado_solicitud_id", "INT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("clave", "VARCHAR(30)", "NN", "-", "UK", "-", "Clave"),
            ("nombre", "VARCHAR(100)", "NN", "-", "-", "-", "Nombre"),
            ("es_terminal", "BIT", "NN", "0", "-", "-", "Estado terminal"),
            ("activa", "BIT", "NN", "1", "-", "-", "Vigencia"),
        ],
    },
    {
        "name": "solicitud",
        "description": "Entidad principal de gestión de solicitudes.",
        "columns": [
            ("solicitud_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("folio", "VARCHAR(30)", "NN", "-", "UK", "-", "Folio único"),
            ("titulo", "VARCHAR(220)", "NN", "-", "-", "-", "Título"),
            ("descripcion", "VARCHAR(3000)", "N", "-", "-", "-", "Descripción"),
            ("area_id", "INT", "NN", "-", "FK", "area.area_id", "Área solicitante"),
            ("sistema_id", "INT", "NN", "-", "FK", "sistema.sistema_id", "Sistema"),
            ("tipo_solicitud_id", "INT", "NN", "-", "FK", "tipo_solicitud.tipo_solicitud_id", "Tipo"),
            ("prioridad_solicitud_id", "INT", "NN", "-", "FK", "prioridad_solicitud.prioridad_solicitud_id", "Prioridad"),
            ("estado_solicitud_id", "INT", "NN", "-", "FK", "estado_solicitud.estado_solicitud_id", "Estado actual"),
            ("creado_por_usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Usuario creador"),
            ("fecha_creacion", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha de creación"),
        ],
    },
    {
        "name": "solicitud_desarrollador",
        "description": "Asignación N:M de desarrolladores a solicitudes.",
        "columns": [
            ("solicitud_desarrollador_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador técnico"),
            ("solicitud_id", "BIGINT", "NN", "-", "FK", "solicitud.solicitud_id", "Solicitud"),
            ("usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Desarrollador"),
            ("rol_en_solicitud", "VARCHAR(40)", "NN", "-", "-", "-", "Rol operativo"),
            ("fecha_asignacion", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha de asignación"),
        ],
    },
    {
        "name": "historial_estado_solicitud",
        "description": "Histórico de cambios de estado de solicitud.",
        "columns": [
            ("historial_estado_solicitud_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("solicitud_id", "BIGINT", "NN", "-", "FK", "solicitud.solicitud_id", "Solicitud"),
            ("estado_solicitud_id", "INT", "NN", "-", "FK", "estado_solicitud.estado_solicitud_id", "Estado"),
            ("usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Usuario que cambia"),
            ("comentario", "VARCHAR(1500)", "N", "-", "-", "-", "Comentario del cambio"),
            ("fecha_cambio", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha de cambio"),
        ],
    },
    {
        "name": "aprobacion_solicitud",
        "description": "Registro de aprobaciones y rechazos.",
        "columns": [
            ("aprobacion_solicitud_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("solicitud_id", "BIGINT", "NN", "-", "FK", "solicitud.solicitud_id", "Solicitud"),
            ("aprobador_usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Aprobador"),
            ("decision", "VARCHAR(20)", "NN", "-", "-", "-", "APROBADA/RECHAZADA"),
            ("comentario", "VARCHAR(1500)", "N", "-", "-", "-", "Comentario"),
            ("fecha_decision", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha de decisión"),
        ],
    },
    {
        "name": "comentario_solicitud",
        "description": "Comentarios asociados a solicitudes.",
        "columns": [
            ("comentario_solicitud_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("solicitud_id", "BIGINT", "NN", "-", "FK", "solicitud.solicitud_id", "Solicitud"),
            ("usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Autor"),
            ("comentario", "VARCHAR(2000)", "NN", "-", "-", "-", "Texto"),
            ("fecha_comentario", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha de comentario"),
        ],
    },
    {
        "name": "adjunto_solicitud",
        "description": "Archivos adjuntos de solicitudes.",
        "columns": [
            ("adjunto_solicitud_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("solicitud_id", "BIGINT", "NN", "-", "FK", "solicitud.solicitud_id", "Solicitud"),
            ("nombre_archivo", "VARCHAR(255)", "NN", "-", "-", "-", "Nombre visible"),
            ("ruta_archivo", "VARCHAR(600)", "NN", "-", "-", "-", "Ruta"),
            ("tamanio_bytes", "BIGINT", "N", "-", "-", "-", "Tamaño"),
            ("subido_por_usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Usuario que sube"),
            ("fecha_subida", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha de carga"),
        ],
    },
    {
        "name": "notificacion",
        "description": "Notificaciones enviadas a usuarios.",
        "columns": [
            ("notificacion_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Destinatario"),
            ("solicitud_id", "BIGINT", "N", "-", "FK", "solicitud.solicitud_id", "Solicitud relacionada"),
            ("titulo", "VARCHAR(220)", "NN", "-", "-", "-", "Título"),
            ("mensaje", "VARCHAR(1500)", "NN", "-", "-", "-", "Mensaje"),
            ("leida", "BIT", "NN", "0", "-", "-", "Estado de lectura"),
            ("fecha_envio", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha de envío"),
        ],
    },
    {
        "name": "usuario_credencial",
        "description": "Credenciales y parámetros de seguridad por usuario.",
        "columns": [
            ("usuario_id", "INT", "NN", "-", "PK/FK", "usuario.usuario_id", "Identificador de usuario"),
            ("login_usuario", "VARCHAR(80)", "NN", "-", "UK", "-", "Nombre de acceso"),
            ("password_hash", "VARBINARY(512)", "NN", "-", "-", "-", "Hash"),
            ("password_salt", "VARBINARY(128)", "NN", "-", "-", "-", "Salt"),
            ("algoritmo_hash", "VARCHAR(30)", "NN", "PBKDF2", "-", "-", "Algoritmo"),
            ("iteraciones", "INT", "NN", "100000", "-", "-", "Iteraciones"),
            ("ultimo_acceso", "DATETIME2", "N", "-", "-", "-", "Último acceso"),
            ("intentos_fallidos", "SMALLINT", "NN", "0", "-", "-", "Intentos fallidos"),
            ("bloqueado_hasta", "DATETIME2", "N", "-", "-", "-", "Bloqueo temporal"),
            ("requiere_cambio_password", "BIT", "NN", "0", "-", "-", "Cambio obligatorio"),
            ("fecha_actualizacion", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Última actualización"),
        ],
    },
    {
        "name": "menu_opcion",
        "description": "Catálogo jerárquico de opciones de menú.",
        "columns": [
            ("menu_opcion_id", "INT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("clave", "VARCHAR(60)", "NN", "-", "UK", "-", "Clave única"),
            ("nombre", "VARCHAR(120)", "NN", "-", "-", "-", "Nombre"),
            ("ruta", "VARCHAR(220)", "NN", "-", "-", "-", "Ruta"),
            ("icono", "VARCHAR(80)", "N", "-", "-", "-", "Ícono"),
            ("orden", "INT", "NN", "-", "-", "-", "Orden"),
            ("activa", "BIT", "NN", "1", "-", "-", "Vigencia"),
            ("menu_padre_id", "INT", "N", "-", "FK", "menu_opcion.menu_opcion_id", "Jerarquía"),
        ],
    },
    {
        "name": "rol_menu_opcion",
        "description": "Permisos por rol sobre opciones de menú.",
        "columns": [
            ("rol_menu_opcion_id", "INT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("rol_id", "INT", "NN", "-", "FK", "rol.rol_id", "Rol"),
            ("menu_opcion_id", "INT", "NN", "-", "FK", "menu_opcion.menu_opcion_id", "Opción"),
            ("puede_ver", "BIT", "NN", "1", "-", "-", "Permiso ver"),
            ("puede_editar", "BIT", "NN", "0", "-", "-", "Permiso editar"),
            ("puede_aprobar", "BIT", "NN", "0", "-", "-", "Permiso aprobar"),
        ],
    },
    {
        "name": "tipo_modificacion",
        "description": "Catálogo de tipos de modificación.",
        "columns": [
            ("tipo_modificacion_id", "INT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("clave", "VARCHAR(30)", "NN", "-", "UK", "-", "Clave"),
            ("nombre", "VARCHAR(80)", "NN", "-", "-", "-", "Nombre"),
            ("activa", "BIT", "NN", "1", "-", "-", "Vigencia"),
        ],
    },
    {
        "name": "solicitud_modificacion",
        "description": "Detalle especializado de solicitudes de modificación.",
        "columns": [
            ("solicitud_modificacion_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("solicitud_id", "BIGINT", "NN", "-", "UK/FK", "solicitud.solicitud_id", "Solicitud base"),
            ("tipo_modificacion_id", "INT", "NN", "-", "FK", "tipo_modificacion.tipo_modificacion_id", "Subtipo"),
            ("sistema_version_actual", "VARCHAR(80)", "N", "-", "-", "-", "Versión actual"),
            ("modulo_afectado", "VARCHAR(200)", "N", "-", "-", "-", "Módulo"),
            ("impacto_tecnico", "VARCHAR(1500)", "N", "-", "-", "-", "Impacto"),
            ("justificacion", "VARCHAR(2000)", "N", "-", "-", "-", "Justificación"),
            ("fecha_registro", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha registro"),
        ],
    },
    {
        "name": "solicitud_requerimiento_tecnico",
        "description": "Especificación técnica de una solicitud.",
        "columns": [
            ("solicitud_requerimiento_tecnico_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("solicitud_id", "BIGINT", "NN", "-", "UK/FK", "solicitud.solicitud_id", "Solicitud base"),
            ("arquitectura_propuesta", "VARCHAR(2000)", "N", "-", "-", "-", "Arquitectura"),
            ("alcance_tecnico", "VARCHAR(2000)", "N", "-", "-", "-", "Alcance"),
            ("dependencias", "VARCHAR(2000)", "N", "-", "-", "-", "Dependencias"),
            ("criterios_aceptacion", "VARCHAR(2000)", "N", "-", "-", "-", "Criterios"),
            ("riesgos", "VARCHAR(2000)", "N", "-", "-", "-", "Riesgos"),
            ("plan_pruebas", "VARCHAR(2000)", "N", "-", "-", "-", "Plan pruebas"),
            ("observaciones", "VARCHAR(2000)", "N", "-", "-", "-", "Observaciones"),
            ("elaborado_por_usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Autor"),
            ("fecha_elaboracion", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha elaboración"),
        ],
    },
    {
        "name": "tarea_desarrollo",
        "description": "Tareas operativas derivadas de solicitudes.",
        "columns": [
            ("tarea_desarrollo_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("solicitud_id", "BIGINT", "NN", "-", "FK", "solicitud.solicitud_id", "Solicitud origen"),
            ("titulo", "VARCHAR(220)", "NN", "-", "-", "-", "Título"),
            ("descripcion", "VARCHAR(2000)", "N", "-", "-", "-", "Descripción"),
            ("prioridad_solicitud_id", "INT", "NN", "-", "FK", "prioridad_solicitud.prioridad_solicitud_id", "Prioridad"),
            ("estado_tarea", "VARCHAR(30)", "NN", "-", "-", "-", "Estado"),
            ("fecha_inicio", "DATETIME2", "N", "-", "-", "-", "Inicio"),
            ("fecha_compromiso", "DATETIME2", "N", "-", "-", "-", "Compromiso"),
            ("fecha_cierre", "DATETIME2", "N", "-", "-", "-", "Cierre"),
            ("creado_por_usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Creador"),
            ("activo", "BIT", "NN", "1", "-", "-", "Vigencia"),
        ],
    },
    {
        "name": "tarea_desarrollo_asignacion",
        "description": "Asignaciones de usuarios a tareas de desarrollo.",
        "columns": [
            ("tarea_desarrollo_asignacion_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("tarea_desarrollo_id", "BIGINT", "NN", "-", "FK", "tarea_desarrollo.tarea_desarrollo_id", "Tarea"),
            ("usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Usuario"),
            ("rol_tecnico", "VARCHAR(30)", "NN", "-", "-", "-", "RESPONSABLE/PARTICIPANTE"),
            ("activa", "BIT", "NN", "1", "-", "-", "Asignación vigente"),
            ("fecha_asignacion", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha asignación"),
            ("fecha_fin", "DATETIME2", "N", "-", "-", "-", "Fecha fin"),
        ],
    },
    {
        "name": "actividad_reciente",
        "description": "Bitácora de actividad para tableros de trabajo.",
        "columns": [
            ("actividad_reciente_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("solicitud_id", "BIGINT", "N", "-", "FK", "solicitud.solicitud_id", "Solicitud"),
            ("tarea_desarrollo_id", "BIGINT", "N", "-", "FK", "tarea_desarrollo.tarea_desarrollo_id", "Tarea"),
            ("usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Actor"),
            ("tipo_actividad", "VARCHAR(40)", "NN", "-", "-", "-", "Tipo"),
            ("titulo", "VARCHAR(220)", "NN", "-", "-", "-", "Título"),
            ("detalle", "VARCHAR(2000)", "N", "-", "-", "-", "Detalle"),
            ("fecha_actividad", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha"),
            ("visible_para_pm", "BIT", "NN", "1", "-", "-", "Visibilidad PM"),
            ("visible_para_desarrollador", "BIT", "NN", "1", "-", "-", "Visibilidad desarrollador"),
        ],
    },
    {
        "name": "proyecto",
        "description": "Entidad de proyecto para agrupar trabajo.",
        "columns": [
            ("proyecto_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("clave", "VARCHAR(40)", "NN", "-", "UK", "-", "Clave única"),
            ("nombre", "VARCHAR(200)", "NN", "-", "-", "-", "Nombre"),
            ("descripcion", "VARCHAR(1500)", "N", "-", "-", "-", "Descripción"),
            ("estado_proyecto", "VARCHAR(30)", "NN", "-", "-", "-", "Estado"),
            ("fecha_inicio", "DATE", "N", "-", "-", "-", "Inicio"),
            ("fecha_fin_planeada", "DATE", "N", "-", "-", "-", "Fin planeado"),
            ("fecha_fin_real", "DATE", "N", "-", "-", "-", "Fin real"),
            ("pm_usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "PM"),
            ("activo", "BIT", "NN", "1", "-", "-", "Vigencia"),
        ],
    },
    {
        "name": "proyecto_solicitud",
        "description": "Relación N:M entre proyectos y solicitudes.",
        "columns": [
            ("proyecto_solicitud_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("proyecto_id", "BIGINT", "NN", "-", "FK", "proyecto.proyecto_id", "Proyecto"),
            ("solicitud_id", "BIGINT", "NN", "-", "FK", "solicitud.solicitud_id", "Solicitud"),
            ("fecha_vinculacion", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha de vínculo"),
        ],
    },
    {
        "name": "proyecto_miembro",
        "description": "Participantes y roles por proyecto.",
        "columns": [
            ("proyecto_miembro_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("proyecto_id", "BIGINT", "NN", "-", "FK", "proyecto.proyecto_id", "Proyecto"),
            ("usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Usuario"),
            ("rol_en_proyecto", "VARCHAR(40)", "NN", "-", "-", "-", "Rol operativo"),
            ("carga_estimada_pct", "DECIMAL(5,2)", "N", "-", "-", "-", "Carga estimada"),
            ("activo", "BIT", "NN", "1", "-", "-", "Vigencia"),
            ("fecha_inicio", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Inicio"),
            ("fecha_fin", "DATETIME2", "N", "-", "-", "-", "Fin"),
        ],
    },
    {
        "name": "documento_proyecto",
        "description": "Documentos asociados a proyectos y solicitudes.",
        "columns": [
            ("documento_proyecto_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("proyecto_id", "BIGINT", "NN", "-", "FK", "proyecto.proyecto_id", "Proyecto"),
            ("solicitud_id", "BIGINT", "N", "-", "FK", "solicitud.solicitud_id", "Solicitud opcional"),
            ("nombre_documento", "VARCHAR(255)", "NN", "-", "-", "-", "Nombre"),
            ("tipo_documento", "VARCHAR(50)", "NN", "-", "-", "-", "Tipo"),
            ("version_documento", "VARCHAR(40)", "N", "-", "-", "-", "Versión"),
            ("ruta_archivo", "VARCHAR(600)", "NN", "-", "-", "-", "Ruta"),
            ("tamanio_bytes", "BIGINT", "N", "-", "-", "-", "Tamaño"),
            ("subido_por_usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Usuario que sube"),
            ("fecha_subida", "DATETIME2", "NN", "SYSDATETIME()", "-", "-", "Fecha de subida"),
            ("visible_para_participantes", "BIT", "NN", "1", "-", "-", "Visibilidad"),
        ],
    },
    {
        "name": "evento_calendario",
        "description": "Eventos de calendario para actividades y reuniones.",
        "columns": [
            ("evento_calendario_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("proyecto_id", "BIGINT", "N", "-", "FK", "proyecto.proyecto_id", "Proyecto"),
            ("solicitud_id", "BIGINT", "N", "-", "FK", "solicitud.solicitud_id", "Solicitud"),
            ("titulo", "VARCHAR(220)", "NN", "-", "-", "-", "Título"),
            ("descripcion", "VARCHAR(2000)", "N", "-", "-", "-", "Descripción"),
            ("tipo_evento", "VARCHAR(40)", "NN", "-", "-", "-", "Tipo de evento"),
            ("fecha_inicio", "DATETIME2", "NN", "-", "-", "-", "Inicio"),
            ("fecha_fin", "DATETIME2", "N", "-", "-", "-", "Fin"),
            ("creado_por_usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Creador"),
            ("es_todo_dia", "BIT", "NN", "0", "-", "-", "Indicador todo el día"),
        ],
    },
    {
        "name": "evento_participante",
        "description": "Participantes invitados a eventos de calendario.",
        "columns": [
            ("evento_participante_id", "BIGINT IDENTITY(1,1)", "NN", "IDENTITY", "PK", "-", "Identificador"),
            ("evento_calendario_id", "BIGINT", "NN", "-", "FK", "evento_calendario.evento_calendario_id", "Evento"),
            ("usuario_id", "INT", "NN", "-", "FK", "usuario.usuario_id", "Participante"),
            ("asistencia", "VARCHAR(20)", "NN", "PENDIENTE", "-", "-", "Estado de asistencia"),
        ],
    },
]

wb = Workbook()
ws_index = wb.active
ws_index.title = "INDICE"

# Encabezados de índice
index_headers = ["Tabla", "Descripción", "Columnas", "Hoja"]
for col, value in enumerate(index_headers, start=1):
    cell = ws_index.cell(row=1, column=col, value=value)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center")

row_idx = 2
for table in tables:
    ws = wb.create_sheet(title=table["name"])

    ws["A1"] = table["name"]
    ws["A1"].font = Font(size=14, bold=True)

    ws["A2"] = "Descripción"
    ws["A2"].font = Font(bold=True)
    ws["A3"] = table["description"]

    # Encabezados técnicos fila 5
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center")

    # Columnas de la tabla desde fila 6
    data_row = 6
    for col_def in table["columns"]:
        for col_pos, item in enumerate(col_def, start=1):
            ws.cell(row=data_row, column=col_pos, value=item)
        data_row += 1

    # Ajuste de ancho aproximado
    widths = {1: 32, 2: 22, 3: 10, 4: 22, 5: 12, 6: 36, 7: 55}
    for col_idx, width in widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Registro en índice
    ws_index.cell(row=row_idx, column=1, value=table["name"])
    ws_index.cell(row=row_idx, column=2, value=table["description"])
    ws_index.cell(row=row_idx, column=3, value=len(table["columns"]))
    ws_index.cell(row=row_idx, column=4, value=table["name"])
    row_idx += 1

# Ajuste de columnas del índice
ws_index.column_dimensions["A"].width = 35
ws_index.column_dimensions["B"].width = 65
ws_index.column_dimensions["C"].width = 12
ws_index.column_dimensions["D"].width = 35

wb.save(FILE_NAME)
print("éxito")
