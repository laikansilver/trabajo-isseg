import re
from pathlib import Path
from openpyxl import load_workbook

root = Path(r"c:/Users/eduar/Documents/trabajo isseg")
input_file = root / "base de datos" / "16_diccionario_datos_modelo_consolidado.xlsx"
output_file = root / "base de datos" / "descripciones_16.tsv"

wb = load_workbook(input_file, data_only=True, read_only=False)
rows_out = []

pat = re.compile(r"^\s*descripci[oó]n\b\s*:?\s*(.*)$", re.IGNORECASE)

for ws in wb.worksheets:
    if ws.title.strip().lower() == "indice":
        continue

    descripcion = ""
    max_r = ws.max_row or 0
    max_c = ws.max_column or 0

    found = False
    for r in range(1, max_r + 1):
        if found:
            break
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            m = pat.match(s)
            if not m:
                continue

            inline = (m.group(1) or "").strip(" :-\t\n\r")
            if inline:
                descripcion = inline
                found = True
                break

            right = ws.cell(row=r, column=c+1).value if c+1 <= max_c else None
            if right is not None and str(right).strip():
                descripcion = str(right).strip()
                found = True
                break

            down = ws.cell(row=r+1, column=c).value if r+1 <= max_r else None
            if down is not None and str(down).strip():
                descripcion = str(down).strip()
                found = True
                break

            descripcion = ""
            found = True
            break

    descripcion = " ".join(str(descripcion).replace("\t", " ").split()) if descripcion is not None else ""
    rows_out.append((ws.title, descripcion))

with output_file.open("w", encoding="utf-8", newline="") as f:
    for t, d in rows_out:
        f.write(f"{t}\t{d}\n")

for t, d in rows_out:
    print(f"{t}\t{d}")
