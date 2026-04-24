from openpyxl import load_workbook
import json
from pathlib import Path

wb = load_workbook(r'C:\Users\eduar\Documents\trabajo isseg\base de datos\diccionario datos-v2.xlsx', data_only=False)

exclude = {'Historial de versiones', 'INDICE'}
tables = {}

for ws in wb.worksheets:
    if ws.title in exclude:
        continue
    
    tabla_info = {
        'nombre': ws.title,
        'descripcion': '',
        'columnas': []
    }
    
    # Leer nombre y descripción de la tabla
    if ws['B1'].value:
        tabla_info['nombre'] = str(ws['B1'].value).strip()
    if ws['B2'].value:
        tabla_info['descripcion'] = str(ws['B2'].value).strip()
    
    # Leer columnas desde fila 4
    for row_idx in range(4, ws.max_row + 1):
        col_name = ws[f'B{row_idx}'].value
        if not col_name or str(col_name).strip() == '':
            break
        
        col_type = ws[f'C{row_idx}'].value or 'VARCHAR(MAX)'
        is_null = ws[f'D{row_idx}'].value
        default_val = ws[f'E{row_idx}'].value
        col_desc = ws[f'F{row_idx}'].value
        relacion = ws[f'G{row_idx}'].value
        
        tabla_info['columnas'].append({
            'nombre': str(col_name).strip(),
            'tipo': str(col_type).strip(),
            'nullable': str(is_null).strip().lower() in ['sí', 'yes', '1'],
            'default': str(default_val).strip() if default_val else None,
            'descripcion': str(col_desc).strip() if col_desc else None,
            'relacion': str(relacion).strip() if relacion else None
        })
    
    tables[ws.title] = tabla_info

# Mostrar un resumen
for tabla, info in tables.items():
    print(f"\n{tabla} ({len(info['columnas'])} columnas)")
    for col in info['columnas'][:3]:
        print(f"  - {col['nombre']}: {col['tipo']} (NULL: {col['nullable']})")
    if len(info['columnas']) > 3:
        print(f"  ... y {len(info['columnas']) - 3} más")

# Guardar JSON
with open('schema.json', 'w', encoding='utf-8') as f:
    json.dump(tables, f, indent=2, ensure_ascii=False)

print(f"\n✓ Esquema guardado en schema.json ({len(tables)} tablas)")
